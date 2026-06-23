"""
LinkedIn Playwright scraper — scrapes the user's own "Applied Jobs" page.
Credentials are stored encrypted; session cookies are persisted so re-login is rare.
"""
from __future__ import annotations

import asyncio
import json
import re
import time
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.ai.provider import decrypt_api_key, encrypt_api_key
from app.database import get_db
from app import models

router = APIRouter(prefix="/api/sync/linkedin", tags=["sync"])


def _commit_with_retry(db, retries: int = 5, delay: float = 2.0) -> None:
    """Commit with retry on SQLite 'database is locked' errors."""
    for attempt in range(retries):
        try:
            db.commit()
            return
        except Exception as exc:
            if attempt < retries - 1 and "locked" in str(exc).lower():
                db.rollback()
                time.sleep(delay * (attempt + 1))
            else:
                raise

# ── In-memory task state ──────────────────────────────────────────────────────

_state: dict = {
    "status": "idle",      # idle | running | done | error | needs_login | needs_2fa
    "step": "",
    "processed": 0,
    "created": 0,
    "updated": 0,
    "skipped": 0,
    "errors": [],
    "log": [],             # per-application action log
    "pagination_log": [],  # debug: pagination events per category (persists after sync)
    "raw_jobs": [],        # all scraped jobs (debug)
    "started_at": None,
    "finished_at": None,
}

# 2FA handoff: background task polls this; submit-2fa endpoint sets it
_2fa_code_input: Optional[str] = None


def _reset_state():
    global _2fa_code_input
    _2fa_code_input = None
    _state.update({
        "status": "idle",
        "step": "",
        "processed": 0,
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "errors": [],
        "log": [],
        "pagination_log": [],
        "started_at": None,
        "finished_at": None,
    })


# ── Schemas ───────────────────────────────────────────────────────────────────

class LinkedInConfigWrite(BaseModel):
    email: str
    password: str


class LinkedInConfigStatus(BaseModel):
    configured: bool
    email: Optional[str] = None
    has_session: bool = False
    last_sync: Optional[str] = None


# ── Config endpoints ──────────────────────────────────────────────────────────

@router.get("/config", response_model=LinkedInConfigStatus)
def get_config(db: Session = Depends(get_db)):
    cfg = db.query(models.LinkedInSync).first()
    if not cfg:
        return LinkedInConfigStatus(configured=False)
    return LinkedInConfigStatus(
        configured=True,
        email=cfg.email,
        has_session=bool(cfg.session_cookies),
        last_sync=cfg.last_sync.isoformat() if cfg.last_sync else None,
    )


@router.post("/config", response_model=LinkedInConfigStatus)
def save_config(payload: LinkedInConfigWrite, db: Session = Depends(get_db)):
    cfg = db.query(models.LinkedInSync).first()
    if cfg:
        cfg.email = payload.email.strip()
        cfg.password_enc = encrypt_api_key(payload.password)
        cfg.session_cookies = None  # force re-login with new credentials
    else:
        cfg = models.LinkedInSync(
            email=payload.email.strip(),
            password_enc=encrypt_api_key(payload.password),
        )
        db.add(cfg)
    db.commit()
    db.refresh(cfg)
    return LinkedInConfigStatus(
        configured=True,
        email=cfg.email,
        has_session=False,
        last_sync=cfg.last_sync.isoformat() if cfg.last_sync else None,
    )


@router.delete("/config")
def delete_config(db: Session = Depends(get_db)):
    db.query(models.LinkedInSync).delete()
    db.commit()
    _reset_state()
    return {"ok": True}


# ── Status / run endpoints ────────────────────────────────────────────────────

@router.get("/status")
def get_status():
    return dict(_state)


@router.get("/debug-excel")
def debug_excel():
    """Download an Excel with all scraped LI jobs from the last sync run."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    log = _state.get("log") or []
    if not log:
        raise HTTPException(404, "Kein Sync-Log vorhanden — bitte zuerst einen LinkedIn-Sync durchführen")

    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Sync"

    headers = ["#", "LI Job-ID", "Firma (LI)", "Rolle (LI)", "Datum (LI)", "Kategorie (LI)", "Status-Hint (LI)", "Aktion", "Status DB", "Raw Context (debug)"]
    header_fill = PatternFill("solid", fgColor="1E4078")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ACTION_COLORS = {
        "neu":              "C6EFCE",
        "aktualisiert":     "FFEB9C",
        "abgesagt":         "FFC7CE",
        "fehler":           "FF0000",
        "unverändert":      "FFFFFF",
        "zur Überprüfung":  "BDD7EE",
    }

    for row_i, entry in enumerate(log, 2):
        row = [
            row_i - 1,
            entry.get("li_job_id", ""),
            entry.get("firma_li", ""),
            entry.get("rolle_li", ""),
            entry.get("datum_li", ""),
            entry.get("kategorie_li", ""),
            entry.get("status_hint_li", ""),
            entry.get("aktion", ""),
            entry.get("status_db", ""),
            entry.get("_raw_context", ""),
        ]
        fill_color = ACTION_COLORS.get(entry.get("aktion", ""), "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color) if fill_color != "FFFFFF" else None
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            if fill:
                cell.fill = fill

    col_widths = [5, 16, 40, 45, 14, 16, 20, 14, 28, 60]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.freeze_panes = "A2"

    # ── Sheet 2: Kategorien-Übersicht ────────────────────────────────────────
    ws2 = wb.create_sheet("Kategorien")
    ws2.append(["Kategorie (LI)", "Label", "Gefunden"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cat in (_state.get("category_counts") or []):
        ws2.append([cat.get("card_type", ""), cat.get("label", ""), cat.get("count", 0)])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12

    # ── Sheet 3: Pagination-Log ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Pagination-Log")
    ws3.append(["#", "Eintrag"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for i, entry in enumerate((_state.get("pagination_log") or []), 1):
        ws3.append([i, str(entry)])
    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date
    ts = date.today().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="linkedin_sync_debug_{ts}.xlsx"'},
    )


@router.post("/run")
def run_sync(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    cfg = db.query(models.LinkedInSync).first()
    if not cfg:
        raise HTTPException(400, "LinkedIn not configured")
    if _state["status"] == "running":
        raise HTTPException(409, "Sync already running")

    _reset_state()
    _state["status"] = "running"
    _state["started_at"] = datetime.now(timezone.utc).isoformat()
    background_tasks.add_task(_run_sync_task, cfg.id)
    return dict(_state)


@router.post("/clear-session")
def clear_session(db: Session = Depends(get_db)):
    cfg = db.query(models.LinkedInSync).first()
    if cfg:
        cfg.session_cookies = None
        db.commit()
    return {"ok": True}


class TwoFaPayload(BaseModel):
    code: str


@router.post("/submit-2fa")
def submit_two_fa(payload: TwoFaPayload):
    global _2fa_code_input
    if _state["status"] != "needs_2fa":
        raise HTTPException(409, "No 2FA pending")
    _2fa_code_input = payload.code.strip()
    return {"ok": True}


# ── Playwright scraper ────────────────────────────────────────────────────────

LOGIN_URL = "https://www.linkedin.com/login"
_TRACKER  = "https://www.linkedin.com/jobs-tracker/?stage="

# (card_type, label, default_status, max_pages, url)
# Alle Tabs über jobs-tracker/?stage= — vollständige Pagination für alle.
CATEGORIES: list[tuple[str, str, str, int, str]] = [
    ("SAVED",       "Gespeichert",    "prospecting", 99, _TRACKER + "saved"),
    ("IN_PROGRESS", "In Bearbeitung", "applied",     99, _TRACKER + "in-progress"),
    ("APPLIED",     "Beworben",       "applied",     99, _TRACKER + "applied"),
    ("INTERVIEWS",  "Interviews",     "hr",          99, _TRACKER + "interview"),
    ("ARCHIVED",    "Archiviert",     "rejected",    99, _TRACKER + "archived"),
]

# LinkedIn status footer text → override main_status
_STATUS_MAP = {
    "application viewed":          ("hr",           None),
    "bewerbung gesehen":           ("hr",           None),
    "no longer accepting":         ("rejected",     None),
    "stelle nicht mehr verfügbar": ("rejected",     None),
    "offer":                       ("negotiating",  None),
    "angebot":                     ("negotiating",  None),
    "interview":                   ("hr",           "1_scheduled"),
}


def _parse_date(text: str) -> Optional[str]:
    """Parse LinkedIn date strings → YYYY-MM-DD."""
    from datetime import timedelta
    t = text.lower().strip()
    # short form: "6d ago", "2w ago", "3mo ago"
    m = re.search(r"(\d+)\s*d\b", t)
    if m:
        return (datetime.now() - timedelta(days=int(m.group(1)))).strftime("%Y-%m-%d")
    m = re.search(r"(\d+)\s*w\b", t)
    if m:
        return (datetime.now() - timedelta(weeks=int(m.group(1)))).strftime("%Y-%m-%d")
    m = re.search(r"(\d+)\s*mo\b", t)
    if m:
        return (datetime.now() - timedelta(days=int(m.group(1)) * 30)).strftime("%Y-%m-%d")
    # "2m ago" — bare 'm' for months (word boundary ensures 'mo' above takes priority)
    m = re.search(r"(\d+)\s*m\b", t)
    if m:
        return (datetime.now() - timedelta(days=int(m.group(1)) * 30)).strftime("%Y-%m-%d")
    # long form: "3 days ago", "2 weeks ago", "1 month ago"
    m = re.search(r"(\d+)\s+day", t)
    if m:
        return (datetime.now() - timedelta(days=int(m.group(1)))).strftime("%Y-%m-%d")
    m = re.search(r"(\d+)\s+week", t)
    if m:
        return (datetime.now() - timedelta(weeks=int(m.group(1)))).strftime("%Y-%m-%d")
    m = re.search(r"(\d+)\s+month", t)
    if m:
        return (datetime.now() - timedelta(days=int(m.group(1)) * 30)).strftime("%Y-%m-%d")
    # absolute: "1/15/2025"
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", t)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


# LinkedIn uses these selectors depending on version/locale
_EMAIL_SELECTORS = ["input[autocomplete='username']", "#username", "input[name='session_key']", "input[type='email']"]
_PASS_SELECTORS  = ["input[autocomplete='current-password']", "#password", "input[name='session_password']", "input[type='password']"]


async def _find_visible_locator(page, selectors: list[str], timeout_total: int = 10000):
    """Find the first *visible* element among all candidate selectors and their matches."""
    import time
    deadline = time.monotonic() + timeout_total / 1000
    while time.monotonic() < deadline:
        for sel in selectors:
            loc = page.locator(sel)
            try:
                count = await loc.count()
            except Exception:
                continue
            for i in range(count):
                try:
                    if await loc.nth(i).is_visible():
                        return loc.nth(i)
                except Exception:
                    continue
        await asyncio.sleep(0.3)
    return None


_PIN_SELECTORS = [
    "input[name='pin']",
    "input[id*='pin']",
    "input[id*='verification']",
    "input[autocomplete='one-time-code']",
    "input[inputmode='numeric']",
    "input[type='number']",
    "input[type='tel']",
]

_REMEMBER_SELECTORS = [
    "#remember-me-prompt__confirm-btn",
    "button[aria-label*='remember' i]",
    "button[aria-label*='erinnern' i]",
    "button[data-litms-control-urn*='remember']",
]


async def _wait_for_2fa_code(timeout: float = 300.0) -> Optional[str]:
    """Poll the global until the submit-2fa endpoint sets a code."""
    global _2fa_code_input
    import time
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if _2fa_code_input is not None:
            code = _2fa_code_input
            _2fa_code_input = None
            return code
        await asyncio.sleep(0.5)
    return None


def _is_checkpoint_url(url: str) -> bool:
    return "checkpoint" in url or "challenge" in url or "uas/login" in url


async def _handle_2fa_checkpoint(page) -> bool:
    """Wait for 2FA: either push-approval (auto page redirect) or manual code entry."""
    global _2fa_code_input
    _2fa_code_input = None
    _state["status"] = "needs_2fa"
    _state["step"] = (
        "LinkedIn verlangt Bestätigung — "
        "App-Benachrichtigung auf dem Handy bestätigen ODER Code aus E-Mail/SMS eingeben"
    )

    import time
    deadline = time.monotonic() + 300  # 5 min total

    while time.monotonic() < deadline:
        # ── Option A: push-notification approved → page auto-redirected ──
        current_url = page.url
        if not _is_checkpoint_url(current_url):
            _state["status"] = "running"
            _state["step"] = "Anmelden: erfolgreich (App-Bestätigung erkannt)"
            # Accept "Remember this device" if shown
            try:
                remember_loc = await _find_visible_locator(page, _REMEMBER_SELECTORS, timeout_total=2000)
                if remember_loc:
                    await remember_loc.click()
            except Exception:
                pass
            return True

        # ── Option B: user typed a code in the app ──
        if _2fa_code_input is not None:
            code = _2fa_code_input
            _2fa_code_input = None
            _state["status"] = "running"
            _state["step"] = "Anmelden: 2FA-Code eingeben…"

            pin_loc = await _find_visible_locator(page, _PIN_SELECTORS, timeout_total=5000)
            if pin_loc:
                await pin_loc.fill(code)
            else:
                await page.keyboard.type(code)
            await page.keyboard.press("Enter")

            _state["step"] = "Anmelden: warte auf Weiterleitung nach 2FA…"
            try:
                await page.wait_for_url(
                    re.compile(r"linkedin\.com/(feed|checkpoint|jobs|my-items|uas/login)"),
                    timeout=20000,
                )
            except Exception:
                pass

            if _is_checkpoint_url(page.url):
                _state["status"] = "error"
                _state["step"] = "2FA fehlgeschlagen — Code falsch oder abgelaufen?"
                return False

            try:
                remember_loc = await _find_visible_locator(page, _REMEMBER_SELECTORS, timeout_total=2000)
                if remember_loc:
                    await remember_loc.click()
            except Exception:
                pass

            _state["step"] = "Anmelden: erfolgreich (Code bestätigt)"
            return True

        await asyncio.sleep(1)

    _state["status"] = "error"
    _state["step"] = "2FA-Timeout — keine Bestätigung erhalten (5 min)"
    return False


async def _login(page, email: str, password: str) -> bool:
    """Attempt email/password login. Returns True if successful."""
    try:
        _state["step"] = "Anmelden: Login-Seite laden…"
        await page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
        # Give JS a moment to hydrate before checking selectors
        await asyncio.sleep(2)

        _state["step"] = "Anmelden: warte auf Login-Formular…"
        email_loc = await _find_visible_locator(page, _EMAIL_SELECTORS, timeout_total=10000)
        if not email_loc:
            current_url = page.url
            title = await page.title()
            try:
                snippet = await page.evaluate("document.body.innerText.slice(0, 300)")
            except Exception:
                snippet = "(kein Text)"
            _state["errors"].append(
                f"Login-Formular nicht gefunden — URL: {current_url} | Titel: {title} | Inhalt: {snippet}"
            )
            _state["status"] = "needs_login"
            _state["step"] = "LinkedIn zeigt keine Login-Maske — Session zurücksetzen und erneut versuchen"
            return False

        pass_loc = await _find_visible_locator(page, _PASS_SELECTORS, timeout_total=5000)

        _state["step"] = "Anmelden: Zugangsdaten eingeben…"
        await email_loc.fill(email)
        if pass_loc:
            await pass_loc.fill(password)
        else:
            await page.keyboard.press("Tab")
            await page.keyboard.type(password)
        # Submit via Enter (confirmed working; button selectors are unreliable across LI versions)
        _state["step"] = "Anmelden: Submit…"
        await page.keyboard.press("Enter")
        _state["step"] = "Anmelden: warte auf Weiterleitung…"
        await page.wait_for_url(
            re.compile(r"linkedin\.com/(feed|checkpoint|jobs|my-items|uas/login)"),
            timeout=20000,
        )
        if "checkpoint" in page.url or "challenge" in page.url:
            return await _handle_2fa_checkpoint(page)
        _state["step"] = "Anmelden: erfolgreich"
        return True
    except Exception as e:
        _state["errors"].append(f"Login-Fehler: {e}")
        _state["step"] = f"Login fehlgeschlagen: {e}"
        return False


_CONSENT_SELECTORS = [
    "button[action-type=ACCEPT]",
    "button[data-test-modal-close-btn]",
]


def _parse_li_entry(chunk: str) -> Optional[dict]:
    """Parse one text chunk between two 'Add note' delimiters into a job dict."""
    lines = [l.strip() for l in chunk.split("\n") if l.strip()]

    # Find first "Firma · Ort" line
    dot_idx = None
    for i, line in enumerate(lines):
        if "·" in line and len(line) < 300:
            dot_idx = i
            break

    if dot_idx is None or dot_idx == 0:
        return None

    parts = lines[dot_idx].split("·", 1)
    firma = parts[0].strip()
    ort = parts[1].strip() if len(parts) > 1 else ""

    # Skip navigation items (tab pills like "Applied · 10")
    if not firma or len(firma) < 3 or ort.strip().isdigit():
        return None

    title = lines[dot_idx - 1]
    if len(title) < 4:
        return None

    # "Applied X ago" line
    beworben_text = ""
    for line in lines[dot_idx + 1:]:
        if re.match(r"Applied\b", line, re.IGNORECASE):
            beworben_text = line
            break

    # Status hint lines
    hinweis = ""
    _HINT_KW = [
        "not moving forward", "application viewed", "resume downloaded",
        "no longer accepting", "offer", "interview scheduled",
    ]
    for line in lines[dot_idx + 1:]:
        if any(kw in line.lower() for kw in _HINT_KW):
            hinweis = line
            break

    return {
        "title": title,
        "firma": firma,
        "ort": ort,
        "beworben": beworben_text,
        "hinweis": hinweis,
    }


async def _accept_consent(page) -> None:
    """Dismiss LinkedIn cookie consent banner if present."""
    for sel in _CONSENT_SELECTORS:
        loc = page.locator(sel)
        try:
            if await loc.count() > 0:
                await loc.first.click()
                await asyncio.sleep(2)
                return
        except Exception:
            continue


async def _scrape_category(page, card_type: str, default_status: str, seen_ids: set[str], max_pages: int = 99, url: str = "") -> list[dict]:
    """Read one LinkedIn job-tracker tab via page text and 'Add note' delimiters."""
    if not url:
        url = _TRACKER + card_type.lower()
    jobs: list[dict] = []
    seen_keys: set[str] = set()
    pages_navigated = 0

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    except Exception:
        return []

    if "login" in page.url or "authwall" in page.url:
        return []

    await _accept_consent(page)
    await asyncio.sleep(3)

    # Save page HTML for offline debugging
    try:
        html = await page.content()
        import pathlib
        pathlib.Path(f"/tmp/linkedin_capture_{card_type}.html").write_text(html, encoding="utf-8")
    except Exception:
        pass

    while True:
        text = await page.inner_text("body")

        # Split on "Add note" — each chunk between two delimiters is one job entry
        chunks = re.split(r"\bAdd\s+note\b", text, flags=re.IGNORECASE)
        _state["pagination_log"].append(
            f"[{card_type}] p{pages_navigated + 1}: {len(chunks) - 1} Einträge im Text"
        )

        for chunk in chunks[:-1]:
            entry = _parse_li_entry(chunk)
            if not entry:
                continue
            dedup_key = f"{entry['firma'].lower().strip()} | {entry['title'].lower().strip()}"
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)

            applied_date = _parse_date(entry["beworben"]) if entry["beworben"] else None

            status_text = (entry["beworben"] + " " + entry["hinweis"]).lower()
            mapped_status: Optional[tuple] = None
            for keyword, mapping in _STATUS_MAP.items():
                if keyword in status_text:
                    mapped_status = mapping
                    break

            jobs.append({
                "id": "",
                "title": entry["title"],
                "company": entry["firma"],
                "ort": entry["ort"],
                "applied_date": applied_date,
                "default_status": default_status,
                "status_hint": mapped_status,
                "hinweis": entry["hinweis"],
                "_raw_context": f"{entry['firma']} · {entry['ort']} | {entry['beworben']} | {entry['hinweis']}",
            })

        if pages_navigated >= max_pages - 1:
            _state["pagination_log"].append(f"[{card_type}] max_pages={max_pages} erreicht, stoppe")
            break

        # Click "Next" for pagination
        clicked_next = False
        loc_a = page.locator('.artdeco-pagination__button--next:not([disabled])').first
        try:
            if await loc_a.count() > 0 and await loc_a.is_visible(timeout=1000):
                await loc_a.scroll_into_view_if_needed(timeout=3000)
                await loc_a.click(timeout=5000)
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    await asyncio.sleep(3)
                pages_navigated += 1
                clicked_next = True
                _state["pagination_log"].append(
                    f"[{card_type}] p{pages_navigated + 1}: Next geklickt, jobs_bisher={len(jobs)}"
                )
        except Exception as e:
            _state["pagination_log"].append(f"[{card_type}] Next-Fehler: {e}")

        if not clicked_next:
            break

    return jobs


def _find_or_create_application(db: Session, job: dict) -> tuple[models.Application, bool]:
    """Match job to existing application or create new. Returns (app, created)."""
    li_job_id = job.get("id", "")

    clean_title = job.get("title", "")

    def _needs_rolle_cleanup(rolle: str) -> bool:
        """True if the stored rolle still contains raw LinkedIn card noise."""
        low = rolle.lower()
        return any(s in low for s in ("applied", "posted", "notes", "add note", "(hybrid)", "(remote)"))

    # 1. Exact match by LinkedIn job ID (prevents duplicates across syncs)
    if li_job_id:
        app = db.query(models.Application).filter(
            models.Application.linkedin_job_id == li_job_id
        ).first()
        if app:
            if clean_title and _needs_rolle_cleanup(app.rolle or ""):
                app.rolle = clean_title
            return app, False

    # 2. Fuzzy match by company + role for jobs without a stored ID yet
    company_lower = job["company"].lower()
    role_lower = job["title"].lower()
    apps = db.query(models.Application).all()
    for app in apps:
        firma_lower = (app.firma or "").lower()
        company_match = (
            company_lower in firma_lower
            or firma_lower in company_lower
            or (app.zielfirma_bei_hh or "").lower() in company_lower
        )
        role_match = (
            role_lower in (app.rolle or "").lower()
            or (app.rolle or "").lower() in role_lower
        )
        if company_match and role_match:
            # Backfill the job ID so future syncs use the fast path
            if li_job_id and not app.linkedin_job_id:
                app.linkedin_job_id = li_job_id
            if clean_title and _needs_rolle_cleanup(app.rolle or ""):
                app.rolle = clean_title
            return app, False

    # 3. Create new application
    initial_status = job.get("default_status", "applied")
    if job.get("status_hint"):
        initial_status = job["status_hint"][0]

    def _to_date(val):
        if val is None:
            return None
        if isinstance(val, date):
            return val
        try:
            return date.fromisoformat(str(val))
        except Exception:
            return None

    applied_date_obj = _to_date(job["applied_date"])

    new_app = models.Application(
        firma=job["company"],
        rolle=job["title"],
        datum_bewerbung=applied_date_obj,
        letztes_update=applied_date_obj,
        quelle="LinkedIn",
        main_status=initial_status,
        abgesagt=(initial_status == "rejected"),
        linkedin_job_id=li_job_id or None,
    )
    db.add(new_app)
    db.flush()

    event_typ = "bewerbung" if initial_status not in ("prospecting", "rejected") else "notiz"
    event_titel = {
        "prospecting": "Gespeichert auf LinkedIn",
        "applied":     "Bewerbung eingereicht",
        "hr":          "Interview",
        "rejected":    "Archiviert / Abgelehnt",
    }.get(initial_status, "Bewerbung eingereicht")

    db.add(models.Event(
        application_id=new_app.id,
        typ=event_typ,
        datum=applied_date_obj,
        titel=event_titel,
        source="linkedin",
    ))
    return new_app, True


def _run_sync_task(cfg_id: int):
    """Blocking wrapper — runs the async scraper from a sync background task."""
    asyncio.run(_async_sync(cfg_id))


async def _async_sync(cfg_id: int):
    from app.database import SessionLocal

    db = SessionLocal()
    try:
        cfg = db.query(models.LinkedInSync).get(cfg_id)
        if not cfg:
            _state["status"] = "error"
            _state["step"] = "Konfiguration nicht gefunden"
            return

        email = cfg.email
        password = decrypt_api_key(cfg.password_enc)

        _state["step"] = "Browser starten…"

        try:
            from playwright.async_api import async_playwright
        except ImportError:
            _state["status"] = "error"
            _state["step"] = "Playwright nicht installiert (docker compose build erforderlich)"
            return

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-infobars",
                    "--window-size=1280,800",
                ],
            )
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1280, "height": 800},
                locale="de-DE",
                extra_http_headers={"Accept-Language": "de-DE,de;q=0.9,en;q=0.8"},
            )
            # Remove webdriver fingerprint
            await context.add_init_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )

            # Restore saved session
            if cfg.session_cookies:
                try:
                    cookies = json.loads(cfg.session_cookies)
                    await context.add_cookies(cookies)
                except Exception:
                    pass

            page = await context.new_page()

            # Check session validity
            _state["step"] = "Session prüfen: öffne LinkedIn…"
            check_url = _TRACKER + "applied"
            try:
                await page.goto(check_url, wait_until="domcontentloaded", timeout=20000)
                _state["step"] = f"Session prüfen: geladen ({page.url[:60]})"
            except Exception as nav_err:
                _state["step"] = f"Session prüfen: Fallback auf Startseite ({nav_err})"
                await page.goto("https://www.linkedin.com", wait_until="domcontentloaded", timeout=20000)

            if "login" in page.url or "authwall" in page.url or "uas/login" in page.url:
                logged_in = await _login(page, email, password)
                if not logged_in:
                    await browser.close()
                    if _state["status"] != "needs_login":
                        _state["status"] = "error"
                    return

            # Save fresh session cookies
            cookies = await context.cookies()
            cfg.session_cookies = json.dumps(cookies)
            _commit_with_retry(db)

            # Scrape all categories; seen_ids is per-category to catch scroll-duplicates,
            # but shared job IDs across categories ARE intentional: ARCHIVED must win
            # over APPLIED even if the same job ID appeared earlier.
            all_jobs_by_id: dict[str, dict] = {}
            category_counts: list[dict] = []
            for card_type, label, default_status, max_pages, cat_url in CATEGORIES:
                _state["step"] = f"Lade Kategorie: {label}…"
                cat_jobs = await _scrape_category(page, card_type, default_status, set(), max_pages=max_pages, url=cat_url)
                for j in cat_jobs:
                    j["_card_type"] = card_type
                    j["_label"] = label
                    # Later categories (higher priority) overwrite earlier ones for the same job ID
                    all_jobs_by_id[j["id"]] = j
                category_counts.append({"card_type": card_type, "label": label, "count": len(cat_jobs)})
                _state["step"] = f"{label}: {len(cat_jobs)} gefunden (gesamt {len(all_jobs_by_id)})"
            all_jobs = list(all_jobs_by_id.values())

            _state["category_counts"] = category_counts
            await browser.close()

        if not all_jobs and _state["status"] != "needs_login":
            _state["step"] = "Keine Jobs gefunden — LinkedIn-Layout evtl. geändert"

        _state["raw_jobs"] = all_jobs  # keep for debug export
        _state["step"] = f"Verarbeite {len(all_jobs)} Einträge…"
        created = updated = skipped = 0
        errors: list[str] = []
        action_log: list[dict] = []
        STATUS_ORDER = ["prospecting", "applied", "hr", "fb", "waiting", "negotiating", "signed", "rejected"]

        for i, job in enumerate(all_jobs):
            _state["processed"] = i + 1
            raw = {
                "li_job_id":      job.get("id", ""),
                "firma_li":       job.get("company", ""),
                "rolle_li":       job.get("title", ""),
                "datum_li":       job.get("applied_date", ""),
                "kategorie_li":   job.get("_label") or job.get("default_status", ""),
                "status_hint_li": str(job.get("status_hint") or ""),
                "_raw_context":   job.get("_raw_context", ""),
            }
            try:
                app, was_created = _find_or_create_application(db, job)
                if was_created:
                    created += 1
                    action_log.append({**raw, "aktion": "neu", "status_db": app.main_status})
                else:
                    target_status = job.get("default_status", "applied")
                    if job.get("status_hint"):
                        target_status = job["status_hint"][0]

                    old_status = app.main_status
                    cur_idx = STATUS_ORDER.index(old_status) if old_status in STATUS_ORDER else 0
                    new_idx = STATUS_ORDER.index(target_status) if target_status in STATUS_ORDER else 0

                    status_changed = (
                        (target_status == "rejected" and old_status != "rejected")
                        or (target_status != "rejected" and new_idx > cur_idx)
                    )

                    if status_changed:
                        # Queue for manual review instead of applying directly
                        li_job_id = job.get("id", "")
                        pm_ext_id = f"linkedin_{li_job_id}__status__{target_status}"
                        already_pending = db.query(models.PendingMatch).filter(
                            models.PendingMatch.source == "linkedin",
                            models.PendingMatch.external_id == pm_ext_id,
                            models.PendingMatch.review_status == "pending",
                        ).first()
                        if not already_pending:
                            sub_hint = job["status_hint"][1] if job.get("status_hint") else None
                            db.add(models.PendingMatch(
                                source="linkedin",
                                external_id=pm_ext_id,
                                confidence=90,
                                event_type="status_change",
                                datum=date.today(),
                                titel=f"Status: {old_status} → {target_status}",
                                suggested_app_id=app.id,
                                suggested_main_status=target_status,
                                suggested_sub_status=sub_hint,
                                status_only=True,
                            ))
                        updated += 1
                        action_log.append({**raw, "aktion": "zur Überprüfung", "status_db": f"{old_status} → {target_status}?"})
                    else:
                        skipped += 1
                        action_log.append({**raw, "aktion": "unverändert", "status_db": old_status})
            except Exception as e:
                errors.append(f"{job.get('company', '?')}: {e}")
                action_log.append({**raw, "aktion": "fehler", "status_db": str(e)})

        cfg.last_sync = datetime.now(timezone.utc)
        _commit_with_retry(db)

        _state.update({
            "status": "done",
            "step": "Fertig",
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
            "log": action_log,
            "finished_at": datetime.now(timezone.utc).isoformat(),
        })

    except Exception as e:
        _state["status"] = "error"
        _state["step"] = f"Fehler: {e}"
        _state["errors"] = [str(e)]
    finally:
        db.close()
