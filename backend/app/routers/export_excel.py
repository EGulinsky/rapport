from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

from app.database import get_db
from app import models
from app.models import EXCEL_EXPORT_MAP
from app.auth.dependencies import get_current_user
from app.routers.applications import apply_ghosting_overrides

router = APIRouter(prefix="/api/export", tags=["export"])

HEADERS = [
    "Firma", "HH?", "Zielfirma", "Rolle", "BesetztvonHH", "Quelle",
    "DatumBewerbung", "LetztesUpdate", "Status", "Ghosting", "Abgesagt",
    "Kommentar", "Gespräch1", "Gespräch2", "Gespräch3", "Gespräch4", "Gespräch5",
]


def fmt_date(d: date | None) -> str:
    return d.strftime("%d.%m.%Y") if d else ""


def flag(b: bool) -> str:
    return "x" if b else ""


@router.get("/excel")
def export_excel(
    show_rejected: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    q = db.query(models.Application)
    if not show_rejected:
        q = q.filter(models.Application.main_status != "rejected")
    apps = q.order_by(models.Application.datum_bewerbung.desc()).all()
    apply_ghosting_overrides(db, apps)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = [30, 6, 25, 35, 25, 15, 14, 14, 40, 9, 9, 40, 35, 35, 35, 35, 35]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Data rows
    for app in apps:
        key = (app.main_status, app.sub_status)
        status_val = EXCEL_EXPORT_MAP.get(key, EXCEL_EXPORT_MAP.get((app.main_status, None), app.main_status))
        ws.append([
            app.firma,
            flag(app.is_headhunter),
            app.zielfirma_bei_hh or "",
            app.rolle,
            app.wurde_besetzt_von or "",
            app.quelle or "",
            fmt_date(app.datum_bewerbung),
            fmt_date(app.letztes_update),
            status_val,
            flag(app.ghosting),
            flag(app.abgesagt),
            app.kommentar or "",
            app.gespraech_1 or "",
            app.gespraech_2 or "",
            app.gespraech_3 or "",
            app.gespraech_4 or "",
            app.gespraech_5 or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%Y-%m-%d")
    filename = f"rapport_export_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
