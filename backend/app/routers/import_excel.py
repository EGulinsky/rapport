from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
from datetime import date
import openpyxl
import io

from app.database import get_db
from app import models, schemas
from app.models import EXCEL_IMPORT_MAP, MAIN_STATUS_LABELS, SUB_STATUS_LABELS

router = APIRouter(prefix="/api/import", tags=["import"])


def parse_date(value) -> date | None:
    """Parse various date formats from Excel."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if hasattr(value, "date"):  # datetime
        return value.date()
    s = str(value).strip()
    if not s or s.lower() in ("none", "nan", ""):
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def cell_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("none", "nan", "") else None


def map_status(excel_status: str | None, abgesagt: bool) -> tuple[str, str | None]:
    if abgesagt:
        return ("rejected", None)
    if not excel_status:
        return ("applied", None)
    return EXCEL_IMPORT_MAP.get(excel_status.strip(), ("applied", None))


@router.post("/excel", response_model=schemas.ImportResult)
async def import_excel(
    file: UploadFile = File(...),
    skip_duplicates: bool = True,
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Nur .xlsx Dateien erlaubt")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel konnte nicht gelesen werden: {e}")

    ws = wb["Tracking"] if "Tracking" in wb.sheetnames else wb.active

    imported = 0
    skipped = 0
    errors = []

    # Build dedup set from existing DB entries
    existing_keys = set()
    if skip_duplicates:
        for a in db.query(models.Application.firma, models.Application.rolle).all():
            existing_keys.add(f"{(a.firma or '').lower().strip()}|{(a.rolle or '').lower().strip()}")

    # Excel columns (1-indexed):
    # 1=Firma 2=HH? 3=Zielfirma 4=Rolle 5=BesetztvonHH 6=Quelle
    # 7=DatumBewerbung 8=LetztesUpdate 9=Status 10=Ghosting 11=Abgesagt
    # 12=Kommentar 13-17=Gespräche

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            firma = cell_str(row[0])
            if not firma:
                continue  # skip empty rows

            hh_flag   = str(row[1]).strip().lower() == "x" if row[1] else False
            zielfirma = cell_str(row[2])
            rolle     = cell_str(row[3]) or "—"
            besetzt   = cell_str(row[4])
            quelle    = cell_str(row[5])
            datum_bew = parse_date(row[6])
            letztes   = parse_date(row[7])
            status_raw= cell_str(row[8])
            ghosting  = str(row[9]).strip().lower() == "x" if row[9] else False
            abgesagt  = str(row[10]).strip().lower() == "x" if row[10] else False
            kommentar = cell_str(row[11])
            g1 = cell_str(row[12]) if len(row) > 12 else None
            g2 = cell_str(row[13]) if len(row) > 13 else None
            g3 = cell_str(row[14]) if len(row) > 14 else None
            g4 = cell_str(row[15]) if len(row) > 15 else None
            g5 = cell_str(row[16]) if len(row) > 16 else None

            dedup_key = f"{firma.lower().strip()}|{rolle.lower().strip()}"
            if skip_duplicates and dedup_key in existing_keys:
                skipped += 1
                continue

            main_status, sub_status = map_status(status_raw, abgesagt)

            app = models.Application(
                firma=firma,
                rolle=rolle,
                main_status=main_status,
                sub_status=sub_status,
                is_headhunter=hh_flag,
                zielfirma_bei_hh=zielfirma,
                quelle=quelle,
                wurde_besetzt_von=besetzt,
                datum_bewerbung=datum_bew,
                letztes_update=letztes,
                kommentar=kommentar,
                gespraech_1=g1,
                gespraech_2=g2,
                gespraech_3=g3,
                gespraech_4=g4,
                gespraech_5=g5,
            )
            db.add(app)
            db.flush()

            ref_date  = datum_bew or letztes or date.today()
            late_date = letztes or datum_bew or date.today()

            db.add(models.Event(
                application_id=app.id,
                typ="bewerbung",
                datum=ref_date,
                titel="Bewerbung eingereicht",
            ))

            if main_status not in ("applied", "prospecting"):
                label = MAIN_STATUS_LABELS.get(main_status, main_status)
                if sub_status:
                    label += f" – {SUB_STATUS_LABELS.get(sub_status, sub_status)}"
                db.add(models.Event(
                    application_id=app.id,
                    typ="status",
                    datum=late_date,
                    titel=label,
                ))

            if kommentar:
                db.add(models.Event(
                    application_id=app.id,
                    typ="notiz",
                    datum=late_date,
                    notiz=kommentar,
                ))

            existing_keys.add(dedup_key)
            imported += 1

        except Exception as e:
            errors.append(f"Zeile {row_idx}: {e}")

    db.commit()

    return schemas.ImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors,
        message=f"{imported} Bewerbungen importiert, {skipped} übersprungen (Duplikate).",
    )
