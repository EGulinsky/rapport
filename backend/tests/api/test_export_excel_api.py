"""L2 API — GET /api/export/excel."""
import io

import openpyxl
import pytest

from tests.factories import application_factory

pytestmark = pytest.mark.api


class TestExportExcel:
    def test_positiv_leere_liste_liefert_valide_datei_mit_headern(self, client):
        resp = client.get("/api/export/excel")

        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Tracking"]
        assert ws.cell(row=1, column=1).value == "Firma"

    def test_positiv_bewerbung_wird_als_zeile_exportiert(self, client, db_session):
        application_factory(
            db_session, firma="Contoso AG", rolle="Backend Engineer",
            main_status="applied", is_headhunter=True, zielfirma_bei_hh="Globex AG",
        )
        db_session.commit()

        resp = client.get("/api/export/excel")

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Tracking"]
        assert ws.cell(row=2, column=1).value == "Contoso AG"
        assert ws.cell(row=2, column=2).value == "x"
        assert ws.cell(row=2, column=3).value == "Globex AG"

    def test_positiv_show_rejected_false_filtert_abgesagte_aus(self, client, db_session):
        application_factory(db_session, firma="Abgesagt", main_status="rejected")
        application_factory(db_session, firma="Aktiv", main_status="applied")
        db_session.commit()

        resp = client.get("/api/export/excel", params={"show_rejected": False})

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Tracking"]
        firmen = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "Aktiv" in firmen
        assert "Abgesagt" not in firmen

    def test_positiv_content_disposition_header_gesetzt(self, client):
        resp = client.get("/api/export/excel")
        assert "attachment" in resp.headers["content-disposition"]
        assert ".xlsx" in resp.headers["content-disposition"]
